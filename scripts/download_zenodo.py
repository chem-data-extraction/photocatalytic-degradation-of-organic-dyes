#!/usr/bin/env python3
import os
import sys
import argparse
import requests
import re
import json
import time
from utils.logger import get_logger
from utils.env import load_dotenv
from utils.config import load_config

logger = get_logger("download_zenodo", "logs/download.log")

CONFIG = {}

def parse_args(yaml_config=None, remaining_argv=None):
    if yaml_config is None:
        yaml_config = {}
    
    download_conf = yaml_config.get("stages", {}).get("download", {})
    
    parser = argparse.ArgumentParser(
        description="Clean minimalist script to download tabular datasets from Zenodo."
    )
    parser.add_argument(
        "--config",
        default="config/default.yaml",
        help="Path to YAML configuration file."
    )
    parser.add_argument(
        "--query", "-q",
        type=str,
        default=download_conf.get("query", "photocatalytic degradation dye dataset"),
        help="Search query for Zenodo."
    )
    parser.add_argument(
        "--limit", "-l",
        type=int,
        default=download_conf.get("limit", 10),
        help="Maximum number of records to retrieve."
    )
    parser.add_argument(
        "--filter", "-m",
        type=str,
        choices=["none", "llm", "interactive"],
        default=download_conf.get("filter_mode", "llm"),
        help="Filtering mode to exclude irrelevant datasets."
    )
    return parser.parse_args(remaining_argv)

def download_file(url, filepath):
    """Downloads a file with a clean console progress bar."""
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"    Failed to connect to download URL: {e}")
        return False

    total_size = int(response.headers.get("content-length", 0))
    block_size = 8192
    downloaded = 0

    try:
        with open(filepath, "wb") as f:
            for chunk in response.iter_content(chunk_size=block_size):
                if chunk:
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total_size > 0:
                        percent = (downloaded / total_size) * 100
                        print(f"\r    Progress: {percent:.1f}% ({downloaded / (1024*1024):.2f}/{total_size / (1024*1024):.2f} MB)", end="", flush=True)
                    else:
                        print(f"\r    Progress: {downloaded / (1024*1024):.2f} MB downloaded", end="", flush=True)
        print()  # New line after progress finishes
        return True
    except Exception as e:
        print()  # New line to clear progress bar line
        logger.error(f"    Failed writing file {filepath}: {e}")
        # Clean up partial file on failure
        if os.path.exists(filepath):
            os.remove(filepath)
        return False

def check_relevance_llm(title, description, filenames):
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        logger.warning("GEMINI_API_KEY is not set. Cannot run LLM filtering. Bypassing check.")
        return True, "GEMINI_API_KEY not configured."
        
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        logger.warning("google-genai is not installed. Cannot run LLM filtering. Bypassing check.")
        return True, "google-genai package is not installed."
        
    try:
        client = genai.Client(api_key=api_key)
        
        schema = {
            "type": "OBJECT",
            "properties": {
                "relevant": {"type": "BOOLEAN"},
                "reason": {"type": "STRING"}
            },
            "required": ["relevant", "reason"]
        }
        
        prompt = f"""Analyze the Zenodo record metadata below.
Determine if it contains a tabular dataset of experiments on PHOTOCATALYTIC DYE DEGRADATION.
The dataset is relevant ONLY if it contains experimental runs measuring dye degradation over time (e.g., initial dye concentration, catalyst dosage, light type, irradiation time, degradation efficiency).
It is IRRELEVANT if:
- It only contains material characterization raw data (like XRD patterns, Raman spectra, XPS spectra, EPR, BET isotherms) without systematic degradation kinetic tables.
- It is about a different chemical process (e.g. gas degradation, bromodichloromethane removal, biological cell processes) instead of photocatalytic dye degradation.

Metadata:
Title: {title}
Description: {description[:1000]}
Files: {filenames}
"""
        config = types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=schema,
            system_instruction="You are an expert scientific data classifier. Determine if a Zenodo dataset contains tabular photocatalytic dye degradation experimental data."
        )
        
        download_conf = CONFIG.get("stages", {}).get("download", {})
        llm_conf = download_conf.get("llm", {})
        model_name = llm_conf.get("metadata_model", "gemini-2.5-flash")
        max_retries = llm_conf.get("max_retries", 5)
        retry_delay = llm_conf.get("retry_delay", 5)
        
        for attempt in range(1, max_retries + 1):
            try:
                response = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=config
                )
                res = json.loads(response.text)
                return res.get("relevant", True), res.get("reason", "No reason provided.")
            except Exception as e:
                logger.warning(f"LLM relevance check failed on attempt {attempt}/{max_retries}: {e}")
                if attempt < max_retries:
                    logger.info(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    logger.error(f"All {max_retries} attempts failed for LLM relevance check. Terminating script.")
                    sys.exit(1)
    except Exception as e:
        logger.error(f"Setup or client initialization for LLM relevance check failed: {e}")
        sys.exit(1)

def get_file_preview(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls', '.ods'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            preview = {}
            for sheetname in wb.sheetnames[:8]:  # Limit to 8 sheets
                sheet = wb[sheetname]
                rows = []
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if r_idx > 6:  # Read 6 rows
                        break
                    row_vals = [str(val)[:40] if val is not None else "" for val in row]
                    if any(row_vals):
                        rows.append(row_vals)
                if rows:
                    preview[sheetname] = rows
            return preview
        except Exception as e:
            return {"error": f"Failed to parse Excel: {e}"}
    elif ext in ('.csv', '.tsv'):
        delim = '\t' if ext == '.tsv' else ','
        try:
            import csv
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f, delimiter=delim)
                rows = []
                for r_idx, row in enumerate(reader):
                    if r_idx > 6:
                        break
                    rows.append([str(val)[:40] for val in row])
                return {"csv_data": rows}
        except Exception as e:
            return {"error": f"Failed to parse CSV: {e}"}
    return {"error": "Unsupported file format"}

def fetch_metadata_context(hit, record_dir):
    """
    Looks for related metadata records, downloads any protocol/metadata files,
    and extracts text context to assist LLM mapping.
    """
    import pandas as pd
    related_ids = hit.get("related_identifiers", [])
    if not related_ids:
        related_ids = hit.get("metadata", {}).get("related_identifiers", [])
    context_parts = []
    
    for rel in related_ids:
        identifier = rel.get("identifier", "")
        
        # Check relation type for metadata links
        rel_type = rel.get("relation_type", {})
        relation_id = ""
        if isinstance(rel_type, dict):
            relation_id = str(rel_type.get("id", "")).lower()
        elif isinstance(rel_type, str):
            relation_id = rel_type.lower()
            
        relation = str(rel.get("relation", "")).lower()
        
        if relation_id == "hasmetadata" or relation == "hasmetadata" or "metadata" in relation_id or "metadata" in relation:
            # Extract recid using regex
            match = re.search(r"zenodo\.(\d+)", identifier)
            if match:
                related_recid = match.group(1)
            else:
                match_digits = re.search(r"(\d+)$", identifier)
                if match_digits:
                    related_recid = match_digits.group(1)
                else:
                    continue
            
            logger.info(f"  Found related metadata record: {related_recid}. Fetching context...")
            api_url = f"https://zenodo.org/api/records/{related_recid}"
            
            try:
                response = requests.get(api_url, timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    
                    files_entries = None
                    files_data = data.get("files")
                    if isinstance(files_data, dict):
                        files_entries = files_data.get("entries", {})
                        
                    files_list = []
                    if files_entries:
                        for k, v in files_entries.items():
                            files_list.append({
                                "key": k,
                                "size": v.get("size"),
                                "download_url": v.get("links", {}).get("content")
                            })
                    elif isinstance(files_data, list):
                        for f in files_data:
                            files_list.append({
                                "key": f.get("key") or f.get("filename"),
                                "size": f.get("size"),
                                "download_url": f.get("links", {}).get("self") or f.get("links", {}).get("content")
                            })
                    
                    for f_info in files_list:
                        key = f_info.get("key", "")
                        download_url = f_info.get("download_url")
                        if not key or not download_url:
                            continue
                        
                        key_lower = key.lower()
                        if any(kw in key_lower for kw in ["metadata", "protocol", "readme", "instruction"]) and key_lower.endswith((".xlsx", ".xls", ".txt", ".md")):
                            dest_path = os.path.join(record_dir, key)
                            logger.info(f"    Downloading metadata file {key}...")
                            
                            dl_resp = requests.get(download_url, timeout=15)
                            if dl_resp.status_code == 200:
                                with open(dest_path, "wb") as f_out:
                                    f_out.write(dl_resp.content)
                                
                                if key_lower.endswith((".xlsx", ".xls")):
                                    try:
                                        xl = pd.ExcelFile(dest_path)
                                        for sheet in xl.sheet_names:
                                            sheet_lower = sheet.lower()
                                            if any(kw in sheet_lower for kw in ["protocol", "method", "descriptor", "dictionary", "measurement", "measurements", "data", "info"]):
                                                df_sheet = xl.parse(sheet)
                                                
                                                col_summary = []
                                                for col in df_sheet.columns:
                                                    col_lower = str(col).lower()
                                                    if any(kw in col_lower for kw in ["material", "nm", "sample", "catalyst", "id", "name"]):
                                                        try:
                                                            unique_vals = df_sheet[col].dropna().unique()
                                                            if len(unique_vals) <= 50:
                                                                col_summary.append(f"  Unique values in '{col}': {list(unique_vals)}")
                                                        except Exception:
                                                            pass
                                                col_summary_str = "Column Summaries:\n" + "\n".join(col_summary) + "\n" if col_summary else ""
                                                
                                                num_rows = 100 if any(kw in sheet_lower for kw in ["protocol", "dictionary", "method"]) else 20
                                                sheet_text = df_sheet.head(num_rows).to_string()
                                                context_parts.append(f"Sheet '{sheet}' in {key}:\n{col_summary_str}{sheet_text}\n")
                                    except Exception as ex_xl:
                                        logger.warning(f"    Failed to parse metadata Excel {key}: {ex_xl}")
                                elif key_lower.endswith((".txt", ".md")):
                                    try:
                                        with open(dest_path, "r", encoding="utf-8", errors="ignore") as f_txt:
                                            txt_content = f_txt.read(10000)
                                            context_parts.append(f"File {key}:\n{txt_content}\n")
                                    except Exception as ex_txt:
                                        logger.warning(f"    Failed to read metadata text {key}: {ex_txt}")
            except Exception as e_api:
                logger.warning(f"  Failed to query related metadata record {related_recid}: {e_api}")
                
    return "\n".join(context_parts) if context_parts else None

def check_file_content_relevance_llm(filepath, title, description, metadata_context=None):
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        return True, {}, "GEMINI_API_KEY not configured."
        
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        return True, {}, "google-genai package is not installed."
        
    preview = get_file_preview(filepath)
    if "error" in preview:
        return True, {}, f"Could not read preview: {preview['error']}"
        
    try:
        client = genai.Client(api_key=api_key)
        
        schema = {
            "type": "OBJECT",
            "properties": {
                "relevant": {"type": "BOOLEAN"},
                "reason": {"type": "STRING"},
                "sheet_name": {"type": "STRING"},
                "column_mapping": {
                    "type": "OBJECT",
                    "properties": {
                        "catalyst_formula": {"type": "STRING", "nullable": True},
                        "dye_name": {"type": "STRING", "nullable": True},
                        "initial_dye_conc_value": {"type": "STRING", "nullable": True},
                        "initial_dye_conc_unit": {"type": "STRING", "nullable": True},
                        "catalyst_dosage_value": {"type": "STRING", "nullable": True},
                        "catalyst_dosage_unit": {"type": "STRING", "nullable": True},
                        "light_type": {"type": "STRING", "nullable": True},
                        "time_value": {"type": "STRING", "nullable": True},
                        "time_unit": {"type": "STRING", "nullable": True},
                        "efficiency_value": {"type": "STRING", "nullable": True}
                    },
                    "required": [
                        "catalyst_formula",
                        "dye_name",
                        "initial_dye_conc_value",
                        "initial_dye_conc_unit",
                        "catalyst_dosage_value",
                        "catalyst_dosage_unit",
                        "light_type",
                        "time_value",
                        "time_unit",
                        "efficiency_value"
                    ]
                }
            },
            "required": ["relevant", "reason"]
        }
        
        prompt = f"""Analyze the content preview of the downloaded file '{os.path.basename(filepath)}'.
This file belongs to the Zenodo record: "{title}"
Description snippet: {description[:500]}
"""
        if metadata_context:
            prompt += f"""
Additional Metadata and Experimental Protocol Context:
{metadata_context[:100000]}
"""
        prompt += f"""
Determine if this file (or one of its sheets) is a clean, flat tabular dataset of photocatalytic dye degradation experiments that is suitable for direct database import according to our schema by applying a column mapping.

A file is SUITABLE (relevant: True) ONLY if:
1. It contains a table structure where each row represents a single experimental observation/run, and columns represent distinct variables.
2. It has a clear, unique mapping to our target fields (e.g. one unique column for Time, and one unique column for Degradation Efficiency or Concentration).

A file is UNSUITABLE (relevant: False) and MUST be rejected if:
1. It is a plotting helper sheet with side-by-side repeated columns for different curves (e.g., multiple 'time' and '% removal' columns side-by-side for different catalyst samples).
2. It is dominated by material characterization data (such as XRD patterns, DRS absorbance/reflectance spectra, BET N2 isotherms, FTIR spectra, EPR/XPS intensities, band gap estimation) rather than systematic dye degradation runs over time.
3. It lacks a flat table structure, making it impossible to import via a simple column-to-field mapping.

If the file is SUITABLE, identify the columns that map to the target fields:
- sheet_name: (For Excel files) the sheet name containing the flat table.
- column_mapping: a dictionary mapping each schema field to the corresponding column name in the dataset (or null / constant). Specifically:
  - catalyst_formula: column name for catalyst formula if present.
  - dye_name: column name for dye name if present.
  - initial_dye_conc_value: column name for initial dye concentration value, or if missing as a column, check the metadata/experimental protocol context. Look specifically at the pollutant/dye degradation experiment description (e.g., "used 150 mL of 7 ppm RhB solution"). Avoid general characterization sample preparation ranges (like "10-100 mg/L" meant for DLS/ELS). Convert concentrations to mg/L (1 ppm = 1 mg/L) and specify it as a numeric constant string (e.g. "7.0").
  - initial_dye_conc_unit: column name for initial dye concentration unit, or a constant like 'mg/L' if not in columns but specified in the context.
  - catalyst_dosage_value: column name for catalyst dosage value, or if missing as a column, check the metadata/experimental protocol context. Check both the degradation protocol description and the measurements data sheet for catalyst concentration/dosage references. For example, if it specifies a constant dosage (like "0.1 g/L", "100 mg/L", or "15 mg in 150 mL" which is 0.1 g/L), or if the measurements sheet has materials named like "TiO2:SiO2 0.1 g/L Fotoc" indicating a 0.1 g/L concentration, you MUST set this field to that constant numeric value as a string (e.g. "0.1"). Avoid using characterization sample preparation ranges like "10-100 mg/L".
  - catalyst_dosage_unit: column name for catalyst dosage unit, or a constant like 'g/L' if not in columns but specified in the context.
  - light_type: column name for light condition/type if present.
  - time_value: column name containing the time values.
  - time_unit: column name containing the time unit, or a constant like 'min', 'hours', 's' if not in columns.
  - efficiency_value: column name containing degradation efficiency (0-100%) or concentration ratio.

CRITICAL INSTRUCTION: If a field is not present as a column in the data preview, but its constant experimental value is described in the provided Metadata/Protocol Context, you MUST map it directly to that constant value (e.g. "7.0" or "0.1") instead of null. Do NOT output range strings like "10-100" as a value. Use only a single, specific constant number.

File structure preview (first few rows of each sheet/data):
{json.dumps(preview, indent=2)}
"""
        config = types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=schema,
            system_instruction="You are a strict data validation agent. Determine if a tabular file contains a clean, flat table (not a plotting coordinate sheet with side-by-side repeated columns or raw characterization data) that is suitable for direct database import using column mapping."
        )
        
        download_conf = CONFIG.get("stages", {}).get("download", {})
        llm_conf = download_conf.get("llm", {})
        model_name = llm_conf.get("content_model", "gemini-3.1-flash-lite")
        max_retries = llm_conf.get("max_retries", 5)
        retry_delay = llm_conf.get("retry_delay", 5)
        
        for attempt in range(1, max_retries + 1):
            try:
                response = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=config
                )
                res = json.loads(response.text)
                
                is_relevant = res.get("relevant", True)
                reason = res.get("reason", "No reason provided.")
                
                mapping = {}
                if is_relevant:
                    mapping["sheet_name"] = res.get("sheet_name")
                    mapping["column_mapping"] = res.get("column_mapping", {})
                    
                return is_relevant, mapping, reason
            except Exception as e:
                logger.warning(f"  LLM content relevance check failed on attempt {attempt}/{max_retries}: {e}")
                if attempt < max_retries:
                    logger.info(f"  Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    logger.error(f"  All {max_retries} attempts failed for LLM content relevance check. Terminating script.")
                    sys.exit(1)
    except Exception as e:
        logger.error(f"  Setup or client initialization for LLM content relevance check failed: {e}")
        sys.exit(1)

def ask_user_interactive(idx, total, record_id, title, description, filenames):
    clean_desc = re.sub(r"<[^>]*>", "", description).strip()
    desc_snippet = clean_desc[:200] + "..." if len(clean_desc) > 200 else clean_desc
    
    logger.plain("=" * 60)
    logger.plain(f"[{idx}/{total}] Record ID: {record_id}")
    logger.plain(f"Title: {title}")
    logger.plain(f"Description: {desc_snippet}")
    logger.plain(f"Tabular Files: {filenames}")
    logger.plain("-" * 60)
    
    while True:
        choice = input("Download this dataset? [y]es / [n]o / [a]ll remaining / [q]uit: ").strip().lower()
        if choice in ("y", "yes"):
            return "y"
        elif choice in ("n", "no"):
            return "n"
        elif choice in ("a", "all"):
            return "a"
        elif choice in ("q", "quit"):
            return "q"
        else:
            logger.warning("Invalid choice. Please enter y, n, a, or q.")


def main():
    load_dotenv()
    
    from utils.config import get_config_and_argv
    global CONFIG
    CONFIG, config_path, remaining_argv = get_config_and_argv()
    
    args = parse_args(CONFIG, remaining_argv)
    
    # Propagate output path from config back to args namespace
    download_conf = CONFIG.get("stages", {}).get("download", {})
    args.output = download_conf.get("output_dir", "data/downloaded")
    
    # Target tabular extensions
    allowed_extensions = (".csv", ".xlsx", ".xls", ".tsv", ".ods")
    
    # Zenodo records API endpoint
    api_url = "https://zenodo.org/api/records"
    
    # We restrict the query specifically to resource_type.type:dataset
    # and combine it with the user's search terms.
    search_query = f'resource_type.type:dataset AND ({args.query})'
    
    params = {
        "q": search_query,
        "size": args.limit,
        "status": "published"
    }
    
    logger.info("Querying Zenodo API...")
    logger.info(f"  Search term: {args.query}")
    logger.info(f"  Output directory: {args.output}")
    logger.info(f"  Record limit: {args.limit}")
    logger.info(f"  Filter: {args.filter}\n")
    
    try:
        response = requests.get(api_url, params=params)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"Failed to query Zenodo API: {e}")
        sys.exit(1)
        
    results = response.json()
    hits = results.get("hits", {}).get("hits", [])
    
    if not hits:
        logger.warning("No records found matching the query.")
        sys.exit(0)
        
    logger.success(f"Found {len(hits)} dataset records. Starting download of tabular files...\n")
    
    os.makedirs(args.output, exist_ok=True)
    
    downloaded_records_count = 0
    downloaded_files_count = 0
    
    for idx, hit in enumerate(hits, 1):
        record_id = hit.get("id")
        metadata = hit.get("metadata", {})
        title = metadata.get("title", "Untitled Dataset")
        
        # Get list of files
        files = hit.get("files", [])
        
        # Filter for tabular files
        tabular_files = [
            f for f in files 
            if f.get("key", "").lower().endswith(allowed_extensions)
        ]
        
        if not tabular_files:
            # Skip record if there are no tabular files
            continue
            
        filenames = [f.get("key", "") for f in tabular_files]
        description = metadata.get("description", "")

        is_relevant = True
        skip_reason = ""
        
        if args.filter == "llm":
            is_relevant, skip_reason = check_relevance_llm(title, description, filenames)
        elif args.filter == "interactive":
            action = ask_user_interactive(idx, len(hits), record_id, title, description, filenames)
            if action == "q":
                logger.info("Exiting download script.")
                break
            elif action == "n":
                is_relevant = False
                skip_reason = "Skipped by user."
            elif action == "a":
                args.filter = "none"
                is_relevant = True
            else:
                is_relevant = True
 
        if not is_relevant:
            logger.info(f"[{idx}/{len(hits)}] Skipping Record ID: {record_id} - {title[:50]}...")
            logger.info(f"    Reason: {skip_reason}")
            logger.plain("-" * 50)
            continue
            
        logger.info(f"[{idx}/{len(hits)}] Processing Record ID: {record_id}")
        logger.info(f"    Title: {title}")
        logger.info(f"    Found {len(tabular_files)} tabular file(s).")
        
        # Create record-specific subdirectory to avoid naming conflicts
        record_dir = os.path.join(args.output, f"zenodo_{record_id}")
        os.makedirs(record_dir, exist_ok=True)
        
        # Fetch metadata context (e.g. protocols/dictionaries) from related records if any
        metadata_context = fetch_metadata_context(hit, record_dir)

        record_has_downloads = False
        downloaded_paths = []
        for f_info in tabular_files:
            filename = f_info.get("key")
            download_url = f_info.get("links", {}).get("self") or f_info.get("links", {}).get("content")
            
            if not filename or not download_url:
                continue
                
            dest_path = os.path.join(record_dir, filename)
            logger.info(f"  Downloading {filename}...")
            
            success = download_file(download_url, dest_path)
            if success:
                # Stage 2 validation check on the downloaded file content
                if args.filter == "llm":
                    logger.info("  Inspecting file content for relevance...")
                    is_rel, mapping, reason = check_file_content_relevance_llm(dest_path, title, description, metadata_context)
                    if not is_rel:
                        logger.warning(f"    File {filename} is IRRELEVANT after content inspection: {reason}")
                        logger.info(f"    Deleting {filename}.")
                        try:
                            os.remove(dest_path)
                        except Exception as e:
                            logger.error(f"    Failed to delete file: {e}")
                        continue
                    else:
                        logger.success(f"    File {filename} is RELEVANT. Reason: {reason}")

                        # Save mapping
                        mapping_path = os.path.join(record_dir, f"{filename}_mapping.json")
                        try:
                            with open(mapping_path, "w", encoding="utf-8") as f_map:
                                json.dump(mapping, f_map, indent=2, ensure_ascii=False)
                            logger.success(f"    Saved column mapping to: {mapping_path}")
                        except Exception as e:
                            logger.warning(f"    Failed to save mapping: {e}")
                
                downloaded_paths.append(dest_path)
                downloaded_files_count += 1
                record_has_downloads = True
                
        if record_has_downloads:
            downloaded_records_count += 1
        else:
            # Clean up empty directory if no files were kept
            try:
                os.rmdir(record_dir)
                logger.info(f"  Removed empty directory for record {record_id}")
            except Exception:
                pass
        logger.plain("-" * 50)
        
    logger.success(f"\nCompleted! Downloaded {downloaded_files_count} files from {downloaded_records_count} datasets.")
    logger.success(f"All files saved under: {os.path.abspath(args.output)}")

if __name__ == "__main__":
    main()
