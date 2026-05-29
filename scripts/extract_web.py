#!/usr/bin/env python3
"""
Web Extraction Driver.
Downloads photocatalysis datasets from Zenodo using query parameters,
applies LLM metadata/content filtering, maps dataset columns to target schema,
and writes to data/extracted/web_extracted_records.csv and data/extracted/extraction_log.jsonl.
"""

from __future__ import annotations

import os
import sys
import glob
import re
import json
import time
import argparse
from datetime import datetime, timezone
from pathlib import Path
import requests
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from utils.logger import get_logger
from utils.env import load_dotenv

logger = get_logger("extract_web", "logs/download.log")

# Configuration constants replacing config/default.yaml
DOWNLOAD_QUERY = "photocatalytic degradation dye dataset"
DOWNLOAD_LIMIT = 10
DOWNLOAD_FILTER_MODE = "llm"
DOWNLOAD_OUTPUT_DIR = "data/raw/external"

METADATA_MODEL = "gemini-2.5-flash"
CONTENT_MODEL = "gemini-3.1-flash-lite"
MAX_RETRIES = 5
RETRY_DELAY = 5

# ==============================================================================
# STAGE 1: Zenodo Download & Relevance Check (from download_zenodo.py)
# ==============================================================================
def download_file(url: str, filepath: str) -> bool:
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"    Failed to connect to download URL: {e}")
        return False

    total_size = int(response.headers.get("content-length", 0))
    block_size = 8192
    downloaded = 0

    try:
        with open(filepath, "wb") as f:
            for chunk in response.iter_content(chunk_size=block_size):
                if chunk:
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total_size > 0:
                        percent = (downloaded / total_size) * 100
                        print(f"\r    Progress: {percent:.1f}% ({downloaded / (1024*1024):.2f}/{total_size / (1024*1024):.2f} MB)", end="", flush=True)
                    else:
                        print(f"\r    Progress: {downloaded / (1024*1024):.2f} MB downloaded", end="", flush=True)
        print()
        return True
    except Exception as e:
        print()
        logger.error(f"    Failed writing file {filepath}: {e}")
        if os.path.exists(filepath):
            os.remove(filepath)
        return False

def check_relevance_llm(title: str, description: str, filenames: list[str]) -> tuple[bool, str]:
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        logger.warning("GEMINI_API_KEY is not set. Bypassing check.")
        return True, "GEMINI_API_KEY not configured."
        
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        logger.warning("google-genai is not installed. Bypassing check.")
        return True, "google-genai package is not installed."
        
    try:
        client = genai.Client(api_key=api_key)
        schema = {
            "type": "OBJECT",
            "properties": {
                "relevant": {"type": "BOOLEAN"},
                "reason": {"type": "STRING"}
            },
            "required": ["relevant", "reason"]
        }
        
        prompt = f"""Analyze the Zenodo record metadata below.
Determine if it contains a tabular dataset of experiments on PHOTOCATALYTIC DYE DEGRADATION.
The dataset is relevant ONLY if it contains experimental runs measuring dye degradation over time (e.g., initial dye concentration, catalyst dosage, light type, irradiation time, degradation efficiency).
It is IRRELEVANT if:
- It only contains material characterization raw data (like XRD patterns, Raman spectra, XPS spectra, EPR, BET isotherms) without systematic degradation kinetic tables.
- It is about a different chemical process (e.g. gas degradation, bromodichloromethane removal, biological cell processes) instead of photocatalytic dye degradation.

Metadata:
Title: {title}
Description: {description[:1000]}
Files: {filenames}
"""
        config = types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=schema,
            system_instruction="You are an expert scientific data classifier. Determine if a Zenodo dataset contains tabular photocatalytic dye degradation experimental data."
        )
        
        model_name = METADATA_MODEL
        max_retries = MAX_RETRIES
        retry_delay = RETRY_DELAY
        
        for attempt in range(1, max_retries + 1):
            try:
                response = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=config
                )
                res = json.loads(response.text)
                return res.get("relevant", True), res.get("reason", "No reason provided.")
            except Exception as e:
                logger.warning(f"LLM relevance check failed on attempt {attempt}/{max_retries}: {e}")
                if attempt < max_retries:
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    logger.error(f"All {max_retries} attempts failed for LLM relevance check.")
                    sys.exit(1)
    except Exception as e:
        logger.error(f"Setup or client initialization for LLM relevance check failed: {e}")
        sys.exit(1)

def get_file_preview(filepath: str) -> dict:
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls', '.ods'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            preview = {}
            for sheetname in wb.sheetnames[:8]:
                sheet = wb[sheetname]
                rows = []
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if r_idx > 6:
                        break
                    row_vals = [str(val)[:40] if val is not None else "" for val in row]
                    if any(row_vals):
                        rows.append(row_vals)
                if rows:
                    preview[sheetname] = rows
            return preview
        except Exception as e:
            return {"error": f"Failed to parse Excel: {e}"}
    elif ext in ('.csv', '.tsv'):
        delim = '\t' if ext == '.tsv' else ','
        try:
            import csv
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f, delimiter=delim)
                rows = []
                for r_idx, row in enumerate(reader):
                    if r_idx > 6:
                        break
                    rows.append([str(val)[:40] for val in row])
                return {"csv_data": rows}
        except Exception as e:
            return {"error": f"Failed to parse CSV: {e}"}
    return {"error": "Unsupported file format"}

def fetch_metadata_context(hit: dict, record_dir: str) -> str | None:
    related_ids = hit.get("related_identifiers", [])
    if not related_ids:
        related_ids = hit.get("metadata", {}).get("related_identifiers", [])
    context_parts = []
    
    for rel in related_ids:
        identifier = rel.get("identifier", "")
        rel_type = rel.get("relation_type", {})
        relation_id = ""
        if isinstance(rel_type, dict):
            relation_id = str(rel_type.get("id", "")).lower()
        elif isinstance(rel_type, str):
            relation_id = rel_type.lower()
            
        relation = str(rel.get("relation", "")).lower()
        
        if relation_id == "hasmetadata" or relation == "hasmetadata" or "metadata" in relation_id or "metadata" in relation:
            match = re.search(r"zenodo\.(\d+)", identifier)
            if match:
                related_recid = match.group(1)
            else:
                match_digits = re.search(r"(\d+)$", identifier)
                if match_digits:
                    related_recid = match_digits.group(1)
                else:
                    continue
            
            logger.info(f"  Found related metadata record: {related_recid}. Fetching context...")
            api_url = f"https://zenodo.org/api/records/{related_recid}"
            
            try:
                response = requests.get(api_url, timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    files_entries = None
                    files_data = data.get("files")
                    if isinstance(files_data, dict):
                        files_entries = files_data.get("entries", {})
                        
                    files_list = []
                    if files_entries:
                        for k, v in files_entries.items():
                            files_list.append({
                                "key": k,
                                "size": v.get("size"),
                                "download_url": v.get("links", {}).get("content")
                            })
                    elif isinstance(files_data, list):
                        for f in files_data:
                            files_list.append({
                                "key": f.get("key") or f.get("filename"),
                                "size": f.get("size"),
                                "download_url": f.get("links", {}).get("self") or f.get("links", {}).get("content")
                            })
                    
                    for f_info in files_list:
                        key = f_info.get("key", "")
                        download_url = f_info.get("download_url")
                        if not key or not download_url:
                            continue
                        
                        key_lower = key.lower()
                        if any(kw in key_lower for kw in ["metadata", "protocol", "readme", "instruction"]) and key_lower.endswith((".xlsx", ".xls", ".txt", ".md")):
                            dest_path = os.path.join(record_dir, key)
                            logger.info(f"    Downloading metadata file {key}...")
                            
                            dl_resp = requests.get(download_url, timeout=15)
                            if dl_resp.status_code == 200:
                                with open(dest_path, "wb") as f_out:
                                    f_out.write(dl_resp.content)
                                
                                if key_lower.endswith((".xlsx", ".xls")):
                                    try:
                                        xl = pd.ExcelFile(dest_path)
                                        for sheet in xl.sheet_names:
                                            sheet_lower = sheet.lower()
                                            if any(kw in sheet_lower for kw in ["protocol", "method", "descriptor", "dictionary", "measurement", "measurements", "data", "info"]):
                                                df_sheet = xl.parse(sheet)
                                                col_summary = []
                                                for col in df_sheet.columns:
                                                    col_lower = str(col).lower()
                                                    if any(kw in col_lower for kw in ["material", "nm", "sample", "catalyst", "id", "name"]):
                                                        try:
                                                            unique_vals = df_sheet[col].dropna().unique()
                                                            if len(unique_vals) <= 50:
                                                                 col_summary.append(f"  Unique values in '{col}': {list(unique_vals)}")
                                                        except Exception:
                                                            pass
                                                col_summary_str = "Column Summaries:\n" + "\n".join(col_summary) + "\n" if col_summary else ""
                                                num_rows = 100 if any(kw in sheet_lower for kw in ["protocol", "dictionary", "method"]) else 20
                                                sheet_text = df_sheet.head(num_rows).to_string()
                                                context_parts.append(f"Sheet '{sheet}' in {key}:\n{col_summary_str}{sheet_text}\n")
                                    except Exception as ex_xl:
                                        logger.warning(f"    Failed to parse metadata Excel {key}: {ex_xl}")
                                elif key_lower.endswith((".txt", ".md")):
                                    try:
                                        with open(dest_path, "r", encoding="utf-8", errors="ignore") as f_txt:
                                            txt_content = f_txt.read(10000)
                                            context_parts.append(f"File {key}:\n{txt_content}\n")
                                    except Exception as ex_txt:
                                        logger.warning(f"    Failed to read metadata text {key}: {ex_txt}")
            except Exception as e_api:
                logger.warning(f"  Failed to query related metadata record {related_recid}: {e_api}")
                
    return "\n".join(context_parts) if context_parts else None

def check_file_content_relevance_llm(filepath: str, title: str, description: str, metadata_context: str | None = None) -> tuple[bool, dict, str]:
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        return True, {}, "GEMINI_API_KEY not configured."
        
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        return True, {}, "google-genai package is not installed."
        
    preview = get_file_preview(filepath)
    if "error" in preview:
        return True, {}, f"Could not read preview: {preview['error']}"
        
    try:
        client = genai.Client(api_key=api_key)
        schema = {
            "type": "OBJECT",
            "properties": {
                "relevant": {"type": "BOOLEAN"},
                "reason": {"type": "STRING"},
                "sheet_name": {"type": "STRING"},
                "column_mapping": {
                    "type": "OBJECT",
                    "properties": {
                        "catalyst": {"type": "STRING", "nullable": True},
                        "catalyst_band_gap_ev": {"type": "STRING", "nullable": True},
                        "catalyst_surface_area_m2g": {"type": "STRING", "nullable": True},
                        "catalyst_particle_size_nm": {"type": "STRING", "nullable": True},
                        "dye_name": {"type": "STRING", "nullable": True},
                        "initial_dye_concentration_value": {"type": "STRING", "nullable": True},
                        "initial_dye_concentration_unit": {"type": "STRING", "nullable": True},
                        "catalyst_dosage_value": {"type": "STRING", "nullable": True},
                        "catalyst_dosage_unit": {"type": "STRING", "nullable": True},
                        "light_type": {"type": "STRING", "nullable": True},
                        "irradiation_time_value": {"type": "STRING", "nullable": True},
                        "irradiation_time_unit": {"type": "STRING", "nullable": True},
                        "degradation_efficiency_percent": {"type": "STRING", "nullable": True}
                    },
                    "required": [
                        "catalyst", "catalyst_band_gap_ev", "catalyst_surface_area_m2g", "catalyst_particle_size_nm",
                        "dye_name", "initial_dye_concentration_value", "initial_dye_concentration_unit",
                        "catalyst_dosage_value", "catalyst_dosage_unit", "light_type",
                        "irradiation_time_value", "irradiation_time_unit", "degradation_efficiency_percent"
                    ]
                }
            },
            "required": ["relevant", "reason"]
        }
        
        prompt = f"""Analyze the content preview of the downloaded file '{os.path.basename(filepath)}'.
This file belongs to the Zenodo record: "{title}"
Description snippet: {description[:500]}
"""
        if metadata_context:
            prompt += f"""
Additional Metadata and Experimental Protocol Context:
{metadata_context[:100000]}
"""
        prompt += f"""
Determine if this file (or one of its sheets) is a clean, flat tabular dataset of photocatalytic dye degradation experiments that is suitable for direct database import according to our schema by applying a column mapping.
 
A file is SUITABLE (relevant: True) ONLY if:
1. It contains a table structure where each row represents a single experimental observation/run, and columns represent distinct variables.
2. It has a clear, unique mapping to our target fields.
 
A file is UNSUITABLE (relevant: False) and MUST be rejected if:
1. It is a plotting helper sheet with side-by-side repeated columns for different curves.
2. It is dominated by material characterization data.
3. It lacks a flat table structure.
 
If the file is SUITABLE, identify the columns that map to the target fields:
- sheet_name: (For Excel files) the sheet name containing the flat table.
- column_mapping: a dictionary mapping each schema field to the corresponding column name in the dataset (or null / constant). Specifically:
  - catalyst: column name for catalyst formula or composition if present.
  - catalyst_band_gap_ev: column name or constant value for catalyst band gap in eV if present.
  - catalyst_surface_area_m2g: column name or constant value for catalyst specific surface area in m2/g if present.
  - catalyst_particle_size_nm: column name or constant value for catalyst particle size/diameter in nm if present.
  - dye_name: column name for dye name.
  - initial_dye_concentration_value: column name or constant value as a string (e.g. "7.0") from protocol context.
  - initial_dye_concentration_unit: column name or constant like 'mg/L'.
  - catalyst_dosage_value: column name or constant value like "0.1".
  - catalyst_dosage_unit: column name or constant like 'g/L'.
  - light_type: column name for light condition/type if present.
  - irradiation_time_value: column name containing the time values.
  - irradiation_time_unit: column name or constant like 'min', 'hours', 's'.
  - degradation_efficiency_percent: column name containing degradation efficiency (0-100%) or concentration ratio.

CRITICAL INSTRUCTION: If a field is not present as a column in the data preview, but its constant experimental value is described in the provided Metadata/Protocol Context, you MUST map it directly to that constant value (e.g. "7.0" or "0.1") instead of null. Use only a single, specific constant number.

File structure preview (first few rows of each sheet/data):
{json.dumps(preview, indent=2)}
"""
        config = types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=schema,
            system_instruction="You are a strict data validation agent. Determine if a tabular file contains a clean, flat table that is suitable for direct database import using column mapping."
        )
        
        model_name = CONTENT_MODEL
        max_retries = MAX_RETRIES
        retry_delay = RETRY_DELAY
        
        for attempt in range(1, max_retries + 1):
            try:
                response = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=config
                )
                res = json.loads(response.text)
                is_relevant = res.get("relevant", True)
                reason = res.get("reason", "No reason provided.")
                mapping = {}
                if is_relevant:
                    mapping["sheet_name"] = res.get("sheet_name")
                    mapping["column_mapping"] = res.get("column_mapping", {})
                return is_relevant, mapping, reason
            except Exception as e:
                logger.warning(f"  LLM content relevance check failed on attempt {attempt}/{max_retries}: {e}")
                if attempt < max_retries:
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    logger.error(f"  All {max_retries} attempts failed for LLM content relevance check.")
                    sys.exit(1)
    except Exception as e:
        logger.error(f"  Setup or client initialization for LLM content relevance check failed: {e}")
        sys.exit(1)

def run_download(
    query: str = DOWNLOAD_QUERY,
    limit: int = DOWNLOAD_LIMIT,
    filter_mode: str = DOWNLOAD_FILTER_MODE,
    output_dir: str = DOWNLOAD_OUTPUT_DIR
) -> None:
    
    allowed_extensions = (".csv", ".xlsx", ".xls", ".tsv", ".ods")
    api_url = "https://zenodo.org/api/records"
    search_query = f'resource_type.type:dataset AND ({query})'
    
    params = {
        "q": search_query,
        "size": limit,
        "status": "published"
    }
    
    logger.info("Querying Zenodo API...")
    try:
        response = requests.get(api_url, params=params)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"Failed to query Zenodo API: {e}")
        return
        
    results = response.json()
    hits = results.get("hits", {}).get("hits", [])
    
    if not hits:
        logger.warning("No records found matching the query.")
        return
        
    logger.success(f"Found {len(hits)} dataset records. Starting download...\n")
    os.makedirs(output_dir, exist_ok=True)
    
    for idx, hit in enumerate(hits, 1):
        record_id = hit.get("id")
        metadata = hit.get("metadata", {})
        title = metadata.get("title", "Untitled Dataset")
        files = hit.get("files", [])
        tabular_files = [f for f in files if f.get("key", "").lower().endswith(allowed_extensions)]
        
        if not tabular_files:
            continue
            
        filenames = [f.get("key", "") for f in tabular_files]
        description = metadata.get("description", "")
        is_relevant = True
        skip_reason = ""
        
        if filter_mode == "llm":
            is_relevant, skip_reason = check_relevance_llm(title, description, filenames)
            
        if not is_relevant:
            logger.info(f"[{idx}/{len(hits)}] Skipping Record ID: {record_id} - {title[:50]}...")
            logger.info(f"    Reason: {skip_reason}")
            continue
            
        logger.info(f"[{idx}/{len(hits)}] Processing Record ID: {record_id}")
        record_dir = os.path.join(output_dir, f"zenodo_{record_id}")
        os.makedirs(record_dir, exist_ok=True)
        
        metadata_context = fetch_metadata_context(hit, record_dir)
        record_has_downloads = False
        
        for f_info in tabular_files:
            filename = f_info.get("key")
            download_url = f_info.get("links", {}).get("self") or f_info.get("links", {}).get("content")
            if not filename or not download_url:
                continue
                
            dest_path = os.path.join(record_dir, filename)
            mapping_path = os.path.join(record_dir, f"{filename}_mapping.json")
            if os.path.exists(dest_path) and os.path.exists(mapping_path):
                logger.info(f"  File {filename} and its mapping already exist. Skipping download.")
                record_has_downloads = True
                continue
                
            logger.info(f"  Downloading {filename}...")
            
            success = download_file(download_url, dest_path)
            if success:
                if filter_mode == "llm":
                    logger.info("  Inspecting file content for relevance...")
                    is_rel, mapping, reason = check_file_content_relevance_llm(dest_path, title, description, metadata_context)
                    if not is_rel:
                        logger.warning(f"    File {filename} is IRRELEVANT after content inspection: {reason}")
                        try:
                            os.remove(dest_path)
                        except Exception:
                            pass
                        continue
                    else:
                        logger.success(f"    File {filename} is RELEVANT.")
                        mapping_path = os.path.join(record_dir, f"{filename}_mapping.json")
                        try:
                            with open(mapping_path, "w", encoding="utf-8") as f_map:
                                json.dump(mapping, f_map, indent=2, ensure_ascii=False)
                        except Exception as e:
                            logger.warning(f"    Failed to save mapping: {e}")
                record_has_downloads = True
                
        if not record_has_downloads:
            try:
                os.rmdir(record_dir)
            except Exception:
                pass

# ==============================================================================
# STAGE 2: Parse Downloaded Mapped Data & Consolidate (from merge_extracted.py)
# ==============================================================================
def to_float(val: object) -> float | None:
    if pd.isna(val) or val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def to_str(val: object) -> str | None:
    if pd.isna(val) or val is None or val == "null" or val == "None":
        return None
    val_str = str(val).strip()
    return val_str if val_str else None

def load_schema_columns(schema_path: Path) -> list[str]:
    with schema_path.open(encoding="utf-8") as f:
        schema = json.load(f)
    return [field["name"] for field in schema["fields"]]

def consolidate_web_records(downloaded_dir: Path, output_csv: Path, schema_path: Path) -> None:
    columns = load_schema_columns(schema_path)
    mapping_files = sorted(list(downloaded_dir.glob("**/*_mapping.json")))
    logger.info(f"Consolidating web records from {len(mapping_files)} dataset mapping file(s) in '{downloaded_dir}'.")
    
    records = []
    record_counter = 1
    
    for mapping_file in mapping_files:
        data_file_path = mapping_file.parent / mapping_file.name.replace("_mapping.json", "")
        if not data_file_path.exists():
            logger.warning(f"Data file {data_file_path} not found for mapping {mapping_file}. Skipping.")
            continue

        try:
            with open(mapping_file, "r", encoding="utf-8") as f:
                mapping = json.load(f)
        except Exception as e:
            logger.error(f"Failed to read mapping file {mapping_file}: {e}. Skipping.")
            continue

        sheet_name = mapping.get("sheet_name")
        suffix = data_file_path.suffix.lower()

        try:
            if suffix in ['.xlsx', '.xls', '.ods']:
                if sheet_name and sheet_name != "null":
                    df = pd.read_excel(data_file_path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(data_file_path)
            elif suffix in ['.csv', '.tsv']:
                sep = '\t' if suffix == '.tsv' else ','
                df = pd.read_csv(data_file_path, sep=sep)
            else:
                continue
        except Exception as e:
            logger.error(f"Failed to read data file {data_file_path}: {e}. Skipping.")
            continue

        col_map = mapping.get("column_mapping", {})
        col_map = {k: (None if v == "null" else v) for k, v in col_map.items()}
        source_id = data_file_path.parent.name

        for _, row in df.iterrows():
            record = {field: None for field in columns}
            record["source_id"] = source_id
            record["record_id"] = f"rec_photo_web_{source_id}_{record_counter:05d}"
            record_counter += 1

            for field in columns:
                if field in ("record_id", "source_id"):
                    continue
                mapped_col = col_map.get(field)
                if not mapped_col:
                    continue

                if isinstance(mapped_col, str) and mapped_col in row:
                    val = row[mapped_col]
                    is_numeric = field.endswith("_value") or field.endswith("_percent") or field in [
                        "catalyst_band_gap_ev", "catalyst_surface_area_m2g", "catalyst_particle_size_nm"
                    ]
                    if is_numeric:
                        record[field] = to_float(val)
                    else:
                        record[field] = to_str(val)
                else:
                    if mapped_col is not None:
                        is_numeric = field.endswith("_value") or field.endswith("_percent") or field in [
                            "catalyst_band_gap_ev", "catalyst_surface_area_m2g", "catalyst_particle_size_nm"
                        ]
                        if is_numeric:
                            record[field] = to_float(mapped_col)
                        else:
                            record[field] = to_str(mapped_col)

            records.append(record)

    out_df = pd.DataFrame(records, columns=columns)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_csv(output_csv, index=False)
    logger.success(f"Wrote {len(out_df)} records to {output_csv.relative_to(ROOT)}")

def append_log(log_path: Path, status: str, count: int, issue: str | None = None) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    entry = {
        "timestamp": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "step": "web_extraction",
        "source_id": "web_manifest",
        "status": status,
        "tool": "extract_web.py",
        "output": "data/extracted/web_extracted_records.csv",
        "records_extracted": count
    }
    if issue:
        entry["issue"] = issue
    with log_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry) + "\n")
    logger.success(f"Appended log event to {log_path.relative_to(ROOT)}")

# ==============================================================================
# MAIN ENTRYPOINT
# ==============================================================================
def main() -> None:
    load_dotenv()
    
    parser = argparse.ArgumentParser(description="Web Extraction Orchestrator.")
    parser.add_argument("--query", default=DOWNLOAD_QUERY, help="Query to search Zenodo.")
    parser.add_argument("--limit", type=int, default=DOWNLOAD_LIMIT, help="Maximum number of records to download.")
    parser.add_argument("--filter-mode", default=DOWNLOAD_FILTER_MODE, choices=["none", "llm", "interactive"], help="Zenodo record filter mode.")
    args = parser.parse_args()
    
    # 1. Download
    logger.info("=== Web Extraction: Downloading Zenodo datasets ===")
    run_download(
        query=args.query,
        limit=args.limit,
        filter_mode=args.filter_mode,
        output_dir=DOWNLOAD_OUTPUT_DIR
    )
    
    # 2. Consolidation
    downloaded_dir = ROOT / DOWNLOAD_OUTPUT_DIR
    output_csv = ROOT / "data/extracted/web_extracted_records.csv"
    schema_path = ROOT / "specs/dataset_schema.json"
    log_path = ROOT / "data/extracted/extraction_log.jsonl"
    
    logger.info("=== Web Extraction: Consolidating Mapped Records ===")
    consolidate_web_records(downloaded_dir, output_csv, schema_path)
    
    try:
        df = pd.read_csv(output_csv)
        count = len(df)
    except Exception:
        count = 0
        
    append_log(log_path, "success", count)
    logger.success("Web Extraction completed successfully!")

if __name__ == "__main__":
    main()
